import os
import json
from datetime import datetime
import calendar

# Imports
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# =============================================================================
# Configuration
# =============================================================================
DATA_FOLDER = "."
DATA_FILE_EXTENSION = ".xlsx"
STATE_FILE = "known_employees.json"

# =============================================================================
# Employee Management
# =============================================================================
def load_known_employees():
    if not os.path.exists(STATE_FILE):
        return set()
    try:
        with open(STATE_FILE, 'r') as f:
            return set(json.load(f).get("employees", []))
    except:
        return set()

def save_known_employees(employees):
    with open(STATE_FILE, 'w') as f:
        json.dump({"employees": sorted(list(employees))}, f, indent=2)

def get_current_files():
    if not os.path.isdir(DATA_FOLDER):
        return set()
    return {os.path.splitext(f)[0] for f in os.listdir(DATA_FOLDER)
            if f.lower().endswith(DATA_FILE_EXTENSION) and not f.lower().startswith("effort")}

def initialize_first_run():
    if os.path.exists(STATE_FILE):
        return
    current = get_current_files()
    if not current:
        return
    print("\n🎉 First run detected!")
    print(f"Found {len(current)} employee files:")
    for i, name in enumerate(sorted(current), 1):
        print(f" {i}. {name}{DATA_FILE_EXTENSION}")
    if input("\nAdd all automatically? (Y/n): ").strip().lower() != 'n':
        save_known_employees(current)
        print("✅ Added all employees!")
    input("\nPress Enter to continue...")

def check_for_changes():
    known = load_known_employees()
    current = get_current_files()
    new = current - known
    missing = known - current
    if new:
        print(f"\n🎉 New files found: {', '.join(sorted(new))}")
        if input("Add them? (y/N): ").lower() == 'y':
            known.update(new)
            save_known_employees(known)
            print("✅ Added!")
    if missing:
        print(f"\n😢 Missing files: {', '.join(sorted(missing))}")
        if input("Remove from tracking? (y/N): ").lower() == 'y':
            known.difference_update(missing)
            save_known_employees(known)
            print("🗑️ Removed.")
    if new or missing:
        input("\nPress Enter to continue...")

# =============================================================================
# Data Loading
# =============================================================================
def get_employee_files():
    return sorted(list(load_known_employees() & get_current_files()))

def _extract_study_hours_from_sheet(df_sheet: pd.DataFrame, sheet_name: str = "") -> pd.DataFrame:
    if df_sheet.empty:
        return pd.DataFrame(columns=["Study ID", "Hours", "Sheet"])
   
    data_start = 0
    for idx, row in df_sheet.iterrows():
        row_str = row.astype(str).str.lower()
        if row_str.str.contains("study id", na=False).any():
            data_start = idx + 1
            break
        if row_str.str.contains(r"^study\d+$", regex=True, na=False).any():
            data_start = idx
            break
   
    df_data = df_sheet.iloc[data_start:].reset_index(drop=True)
   
    study_col = None
    for col in df_data.columns:
        if df_data[col].astype(str).str.fullmatch(r"STUDY\d+", na=False).any():
            study_col = col
            break
   
    if study_col is None:
        return pd.DataFrame(columns=["Study ID", "Hours", "Sheet"])
   
    valid = df_data[study_col].astype(str).str.fullmatch(r"STUDY\d+", na=False)
    if not valid.any():
        return pd.DataFrame(columns=["Study ID", "Hours", "Sheet"])
   
    study_ids = df_data.loc[valid, study_col]
    numeric_cols = df_data.select_dtypes(include="number").columns
    hours = df_data.loc[valid, numeric_cols].sum(axis=1)
    hours = pd.to_numeric(hours, errors="coerce").fillna(0)
   
    return pd.DataFrame({"Study ID": study_ids.values, "Hours": hours.values, "Sheet": sheet_name})

def load_employee_data(employee, raw_sheets=False):
    path = os.path.join(DATA_FOLDER, f"{employee}{DATA_FILE_EXTENSION}")
    if not os.path.exists(path):
        return pd.DataFrame() if not raw_sheets else (pd.DataFrame(), 0)
   
    try:
        excel = pd.ExcelFile(path, engine='openpyxl')
        all_data = []
        sheet_count = 0
        for sheet_name in excel.sheet_names:
            if sheet_name.lower() in ["cover", "summary", "notes", "info"]:
                continue
            df_sheet = pd.read_excel(excel, sheet_name=sheet_name, header=None, engine='openpyxl')
            cleaned = _extract_study_hours_from_sheet(df_sheet, sheet_name)
            if not cleaned.empty:
                all_data.append(cleaned)
                sheet_count += 1
        df = pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame()
        return (df, sheet_count) if raw_sheets else df
    except Exception as e:
        print(f"⚠️ Error reading {employee}: {e}")
        return pd.DataFrame() if not raw_sheets else (pd.DataFrame(), 0)

def load_all_data():
    employees = get_employee_files()
    if not employees:
        return pd.DataFrame()
    dfs = []
    for e in employees:
        df = load_employee_data(e)
        if not df.empty:
            df = df.copy()
            df['Employee'] = e
            dfs.append(df)
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

# =============================================================================
# Plot Functions
# =============================================================================
def safe_plot(func):
    def wrapper(df, title_suffix=""):
        if df.empty:
            print("😔 No data available.")
            input("Press Enter to continue...")
            return
        try:
            func(df, title_suffix)
        except Exception as e:
            print(f"⚠️ Plot error: {e}")
            input("Press Enter to continue...")
    return wrapper

@safe_plot
def plot_bar_hours(df, suffix=""):
    grouped = df.groupby("Study ID")["Hours"].sum().sort_values(ascending=False)
    plt.figure(figsize=(12, 6))
    sns.barplot(x=grouped.index, y=grouped.values)
    plt.title(f'Total Hours per Study{suffix}')
    plt.xticks(rotation=45, ha="right")
    plt.ylabel("Hours")
    plt.tight_layout()
    plt.show()

@safe_plot
def plot_pie_proportions(df, suffix=""):
    grouped = df.groupby("Study ID")["Hours"].sum()
    plt.figure(figsize=(9, 9))
    plt.pie(grouped, labels=grouped.index, autopct='%1.1f%%', startangle=90)
    plt.title(f'Proportion by Study{suffix}')
    plt.axis('equal')
    plt.show()

@safe_plot
def plot_bar_employee_hours(df, suffix=""):
    grouped = df.groupby("Employee")["Hours"].sum().sort_values(ascending=False)
    plt.figure(figsize=(10, 6))
    sns.barplot(x=grouped.index, y=grouped.values)
    plt.title(f'Total Hours by Employee{suffix}')
    plt.xticks(rotation=45, ha="right")
    plt.ylabel("Hours")
    plt.tight_layout()
    plt.show()

@safe_plot
def plot_bar_study_hours_horizontal(df, suffix=""):
    grouped = df.groupby("Study ID")["Hours"].sum().sort_values()
    plt.figure(figsize=(10, max(6, len(grouped)*0.5)))
    sns.barplot(x=grouped.values, y=grouped.index, orient='h')
    plt.title(f'Hours by Study (Horizontal){suffix}')
    plt.xlabel("Hours")
    plt.tight_layout()
    plt.show()

@safe_plot
def plot_stacked_bar(df, suffix=""):
    pivot = df.pivot_table(index="Employee", columns="Study ID", values="Hours", aggfunc="sum", fill_value=0)
    pivot.plot(kind="bar", stacked=True, figsize=(12, 7), colormap="tab20")
    plt.title(f'Stacked Hours by Study{suffix}')
    plt.xticks(rotation=45, ha="right")
    plt.ylabel("Hours")
    plt.legend(title="Study ID", bbox_to_anchor=(1.05, 1), loc='upper left')
    plt.tight_layout()
    plt.show()

@safe_plot
def plot_line_hours_by_week(df, suffix=""):
    if "Sheet" not in df.columns or df["Sheet"].isnull().all():
        print("😔 No weekly data.")
        input("Press Enter...")
        return
    weekly = df.groupby("Sheet")["Hours"].sum().reset_index().sort_values("Sheet")
    plt.figure(figsize=(12, 6))
    sns.lineplot(data=weekly, x="Sheet", y="Hours", marker="o")
    plt.title(f'Hours Over Weeks{suffix}')
    plt.xticks(rotation=45)
    plt.ylabel("Hours")
    plt.tight_layout()
    plt.show()

@safe_plot
def plot_heatmap_hours(df, suffix=""):
    pivot = df.pivot_table(index="Employee", columns="Study ID", values="Hours", aggfunc="sum", fill_value=0)
    plt.figure(figsize=(12, 8))
    sns.heatmap(pivot, annot=True, cmap="YlOrRd", fmt=".1f")
    plt.title(f'Heatmap: Employee vs Study{suffix}')
    plt.tight_layout()
    plt.show()

@safe_plot
def plot_boxplot_hours(df, suffix=""):
    plt.figure(figsize=(12, 6))
    sns.boxplot(data=df, x="Study ID", y="Hours")
    plt.title(f'Boxplot: Distribution{suffix}')
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

@safe_plot
def plot_violin_hours(df, suffix=""):
    plt.figure(figsize=(12, 6))
    sns.violinplot(data=df, x="Study ID", y="Hours")
    plt.title(f'Violin Plot: Distribution{suffix}')
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

@safe_plot
def plot_scatter_hours(df, suffix=""):
    if "Sheet" not in df.columns or df["Sheet"].isnull().all():
        print("😔 No weekly data.")
        input("Press Enter...")
        return
    plt.figure(figsize=(12, 6))
    sns.scatterplot(data=df, x="Sheet", y="Hours", hue="Study ID", s=80)
    plt.title(f'Scatter: Hours by Week{suffix}')
    plt.xticks(rotation=45)
    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
    plt.tight_layout()
    plt.show()

@safe_plot
def plot_area_chart(df, suffix=""):
    if "Sheet" not in df.columns or df["Sheet"].isnull().all():
        print("😔 No weekly data.")
        input("Press Enter...")
        return
    pivot = df.pivot_table(index="Sheet", columns="Study ID", values="Hours", aggfunc="sum", fill_value=0).sort_index()
    pivot.plot(kind="area", stacked=False, figsize=(12, 6), alpha=0.7)
    plt.title(f'Area Chart: Hours Over Time{suffix}')
    plt.xticks(rotation=45)
    plt.ylabel("Hours")
    plt.tight_layout()
    plt.show()

@safe_plot
def plot_stacked_area_chart(df, suffix=""):
    if "Sheet" not in df.columns or df["Sheet"].isnull().all():
        print("😔 No weekly data.")
        input("Press Enter...")
        return
    pivot = df.pivot_table(index="Sheet", columns="Study ID", values="Hours", aggfunc="sum", fill_value=0).sort_index()
    pivot.plot(kind="area", stacked=True, figsize=(12, 6), colormap="Set3")
    plt.title(f'Stacked Area: Hours Over Time{suffix}')
    plt.xticks(rotation=45)
    plt.ylabel("Hours")
    plt.tight_layout()
    plt.show()

# =============================================================================
# Helper: Offer to generate report if no CSV exists
# =============================================================================
def ensure_report_exists():
    csv_files = [f for f in os.listdir(DATA_FOLDER) if f.startswith("EffortData_") and f.endswith(".csv")]
    if csv_files:
        return True
    print("\n😔 No report found for visualisation.")
    if input("Do you want to generate a report first? (y/N): ").strip().lower() == 'y':
        generate_report()
        return True
    return False

# =============================================================================
# Report Generation - UPDATED FOR DYNAMIC NAMING
# =============================================================================
def generate_report():
    employees = get_employee_files()
    if not employees:
        print("😅 No tracked employees!")
        input("Press Enter...")
        return
  
    print("\n📄 Generate Report")
    print("1. All employees")
    print("2. Selected employees")
    choice = input("Choose (1 or 2): ").strip()
  
    if choice == '1':
        selected = employees
    elif choice == '2':
        print("\nSelect employees (space-separated numbers):")
        for i, e in enumerate(sorted(employees), 1):
            print(f" {i}. {e}")
        try:
            nums = list(map(int, input("Numbers: ").split()))
            selected = [sorted(employees)[i-1] for i in nums if 1 <= i <= len(employees)]
            if not selected:
                print("No employees selected.")
                input("Press Enter...")
                return
        except:
            print("Invalid input.")
            input("Press Enter...")
            return
    else:
        print("Invalid choice.")
        input("Press Enter...")
        return
  
    rdate = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Dynamic naming: single employee → use their name, otherwise "Effort"
    if len(selected) == 1:
        prefix = selected[0]
    else:
        prefix = "Effort"
    
    excel_path = f"{prefix}Files_{rdate}.xlsx"
    csv_path = f"{prefix}Data_{rdate}.csv"
    
    all_details = []
    try:
        wb = Workbook()
        wb.remove(wb.active)
        for emp in selected:
            df = load_employee_data(emp)
            if df.empty:
                ws = wb.create_sheet(title=emp[:31])
                ws.append(["No data found or file unreadable"])
                continue
          
            df_copy = df.copy()
            df_copy["Employee"] = emp
            all_details.append(df_copy)
          
            summary = df.groupby("Study ID")["Hours"].sum().reset_index()
            total = summary["Hours"].sum()
            if total == 0:
                ws = wb.create_sheet(title=emp[:31])
                ws.append(["No hours recorded"])
                continue
          
            summary["Total Hours"] = total
            summary["%"] = (summary["Hours"] / total * 100).round(2).astype(str) + "%"
            summary = summary.rename(columns={"Study ID": "Account"})[["Account", "Hours", "Total Hours", "%"]]
          
            ws = wb.create_sheet(title=emp[:31])
            for r in dataframe_to_rows(summary, index=False, header=True):
                ws.append(r)
      
        if len(selected) > 1 and all_details:
            full = pd.concat(all_details)
            pivot = full.pivot_table(index="Study ID", columns="Employee", values="Hours", aggfunc="sum", fill_value=0)
            pivot["Grand Total"] = pivot.sum(axis=1)
            pivot.loc["Total per Employee"] = pivot.sum()
            pivot = pivot.reset_index().rename(columns={"Study ID": "Account"})
            ws = wb.create_sheet("Overall Summary")
            for r in dataframe_to_rows(pivot, index=False, header=True):
                ws.append(r)
      
        wb.save(excel_path)
        print(f"\n🎉 Excel report saved: {excel_path}")
        if all_details:
            final_csv = pd.concat(all_details)[["Employee", "Study ID", "Hours", "Sheet"]]
            final_csv.to_csv(csv_path, index=False)
            print(f"⚡ Fast CSV saved: {csv_path}")
    except Exception as e:
        print(f"\n😱 Report generation failed: {e}")
    input("\nPress Enter to continue...")

# =============================================================================
# Normal Visualisation Menu
# =============================================================================
def visualisation_menu():
    if not ensure_report_exists():
        return
  
    csv_files = sorted([f for f in os.listdir(DATA_FOLDER) if f.startswith("EffortData_") and f.endswith(".csv")], reverse=True)
    df = pd.read_csv(csv_files[0])
  
    while True:
        print("\n" + "🎨"*30)
        print(" VISUALISATION MENU")
        print("🎨"*30)
        print("1. Bar: Hours per Study")
        print("2. Pie: Proportion by Study")
        print("3. Bar: Total Hours by Employee")
        print("4. Horizontal Bar: Hours by Study")
        print("5. Stacked Bar: Studies per Employee")
        print("6. Line: Hours Over Weeks")
        print("7. Heatmap: Employee vs Study")
        print("8. Boxplot: Distribution by Study")
        print("9. Violin Plot: Distribution by Study")
        print("10. Scatter: Hours by Week & Study")
        print("11. Area Chart: Hours Over Time")
        print("12. Stacked Area Chart")
        print("0. Back")
        print("-"*50)
      
        ch = input("Choose (0-12): ").strip()
        if ch == '0':
            return
      
        plots = {
            '1': plot_bar_hours, '2': plot_pie_proportions, '3': plot_bar_employee_hours,
            '4': plot_bar_study_hours_horizontal, '5': plot_stacked_bar, '6': plot_line_hours_by_week,
            '7': plot_heatmap_hours, '8': plot_boxplot_hours, '9': plot_violin_hours,
            '10': plot_scatter_hours, '11': plot_area_chart, '12': plot_stacked_area_chart
        }
      
        if ch in plots:
            plots[ch](df)
        else:
            print("🤔 Invalid choice!")

# =============================================================================
# Advanced Mode
# =============================================================================
def advanced_mode():
    employees = get_employee_files()
    if not employees:
        print("No employees tracked.")
        input("Press Enter...")
        return
  
    while True:
        print("\n🔬 ADVANCED MODE")
        print("1. Employee Sheet Count")
        print("2. Generate Report (Custom)")
        print("3. Visualise Single Employee")
        print("4. Visualise Selected Employees")
        print("5. Filter by Day of Week")
        print("0. Back")
        print("-"*40)
        ch = input("Choose: ").strip()
      
        if ch == '0':
            return
      
        elif ch == '1':
            print("\n📊 Sheet Count per Employee:")
            for e in sorted(employees):
                _, count = load_employee_data(e, raw_sheets=True)
                print(f" {e}: {count} sheets")
            input("\nPress Enter...")
      
        elif ch == '2':
            generate_report()
      
        elif ch in ['3', '4']:
            if not ensure_report_exists():
                raw_df = load_all_data()
                if raw_df.empty:
                    continue
                data = raw_df
            else:
                csv_files = sorted([f for f in os.listdir(DATA_FOLDER) if f.startswith("EffortData_") and f.endswith(".csv")], reverse=True)
                full_data = pd.read_csv(csv_files[0])
                data = full_data
          
            print("\nSelect employees:")
            for i, e in enumerate(sorted(employees), 1):
                print(f" {i}. {e}")
            try:
                nums = list(map(int, input("Numbers (space-separated): ").split()))
                selected = [sorted(employees)[i-1] for i in nums if 1 <= i <= len(employees)]
                if not selected:
                    print("No employees selected.")
                    continue
            except:
                print("Invalid input.")
                continue
          
            filtered = data[data["Employee"].isin(selected)]
            if filtered.empty:
                print("No data for selected employees.")
                continue
          
            print("\nChoose visualisation:")
            print("1. Bar: Hours per Study")
            print("2. Pie: Proportion by Study")
            print("3. Bar: Total Hours by Employee")
            print("4. Horizontal Bar: Hours by Study")
            print("5. Stacked Bar: Studies per Employee")
            print("6. Line: Hours Over Weeks")
            print("7. Heatmap: Employee vs Study")
            print("8. Boxplot: Distribution by Study")
            print("9. Violin Plot: Distribution by Study")
            print("10. Scatter: Hours by Week & Study")
            print("11. Area Chart: Hours Over Time")
            print("12. Stacked Area Chart")
          
            v = input("Choose (1-12): ").strip()
            plots = {
                '1': plot_bar_hours, '2': plot_pie_proportions, '3': plot_bar_employee_hours,
                '4': plot_bar_study_hours_horizontal, '5': plot_stacked_bar, '6': plot_line_hours_by_week,
                '7': plot_heatmap_hours, '8': plot_boxplot_hours, '9': plot_violin_hours,
                '10': plot_scatter_hours, '11': plot_area_chart, '12': plot_stacked_area_chart
            }
            suffix = f" ({', '.join(selected)})" if len(selected) > 1 else f" ({selected[0]})"
            if v in plots:
                plots[v](filtered, suffix)
            else:
                print("Invalid choice.")
      
        elif ch == '5':
            if not ensure_report_exists():
                continue
          
            print("\nDays: 0=Monday ... 6=Sunday")
            try:
                day_num = int(input("Enter day number (0-6): "))
                if not 0 <= day_num <= 6:
                    raise ValueError
                day_name = calendar.day_name[day_num]
            except:
                print("Invalid day.")
                continue
          
            csv_files = sorted([f for f in os.listdir(DATA_FOLDER) if f.startswith("EffortData_") and f.endswith(".csv")], reverse=True)
            df = pd.read_csv(csv_files[0])
          
            def get_weekday(sheet):
                try:
                    for part in str(sheet).split():
                        if len(part) == 10 and '-' in part:
                            return datetime.strptime(part, "%Y-%m-%d").weekday()
                except:
                    pass
                return None
          
            df['Weekday'] = df['Sheet'].apply(get_weekday)
            filtered = df[df['Weekday'] == day_num]
          
            if filtered.empty:
                print(f"No data on {day_name}s.")
            else:
                print(f"📊 {len(filtered)} records on {day_name}s")
                print(f"Total hours: {filtered['Hours'].sum():.1f}")
                if input("Show visualisation? (y/N): ").lower() == 'y':
                    plot_bar_employee_hours(filtered, f" - Only {day_name}s")
      
        input("\nPress Enter to continue...")

# =============================================================================
# Remove Employee
# =============================================================================
def remove_employee():
    employees = sorted(get_employee_files())
    if not employees:
        print("No employees to remove.")
        input("Press Enter...")
        return
  
    print("\n🚪 Remove Employee")
    for i, e in enumerate(employees, 1):
        print(f" {i}. {e}")
  
    try:
        n = int(input("\nEnter number (0 to cancel): "))
        if n == 0:
            return
        if 1 <= n <= len(employees):
            emp = employees[n-1]
            if input(f"Remove {emp}? (y/N): ").lower() == 'y':
                known = load_known_employees()
                known.discard(emp)
                save_known_employees(known)
                print("✅ Removed.")
        else:
            print("Invalid number.")
    except:
        print("Invalid input.")
  
    input("Press Enter...")

# =============================================================================
# Main Menu
# =============================================================================
def main_menu():
    initialize_first_run()
    check_for_changes()
  
    while True:
        employees = sorted(get_employee_files())
        print("\n" + "🔥"*30)
        print(" ACCIDENTAL DRUG-RELATED DEATH ANALYSIS SYSTEM")
        print("🔥"*30)
        print(f"Active employees: {len(employees)} → {', '.join(employees) or 'None yet'}")
        print("\n🌟 MAIN MENU 🌟")
        print("1. 🚪 Remove employee")
        print("2. 📄 Generate Excel Report")
        print("3. 🎨 Visualisations")
        print("4. 🔬 Advanced Mode")
        print("0. 🚪 Exit")
        print("-"*60)
      
        choice = input("Choose (0-4): ").strip()
      
        if choice == '0':
            print("\n👋 Goodbye! Thanks for using the system.\n")
            break
        elif choice == '1':
            remove_employee()
        elif choice == '2':
            generate_report()
        elif choice == '3':
            visualisation_menu()
        elif choice == '4':
            advanced_mode()
        else:
            print("🤨 Invalid option! Try again.")

# =============================================================================
# Start
# =============================================================================
if __name__ == "__main__":
    print("🚀 Starting Accidental Drug-Related Death Analysis System...\n")
    main_menu()